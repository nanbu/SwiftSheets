"""Cross-checks SwiftSheets against openpyxl in both directions with one shared workbook definition.

    python3 Tests/OpenpyxlParity/verify_with_openpyxl.py    # from the repository root, any Python with openpyxl

1. `swift test` writes swiftsheets.xlsx (OpenpyxlInteropTests.writesVerificationWorkbook); openpyxl loads it and
   checks every value, style, dimension, merge, link and name.
2. openpyxl writes openpyxl.xlsx with the same content; `swift test` loads it and checks the same expectations
   (OpenpyxlInteropTests.readsVerificationWorkbook).
Exit 0 when both directions pass. Requires a Python with openpyxl (the Stream web venv has it)."""
import datetime as dt
import json
import os
import pathlib
import subprocess
import sys
import tempfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

PKG = pathlib.Path(__file__).resolve().parents[2]
workdir = pathlib.Path(tempfile.mkdtemp(prefix="swiftsheets-interop-"))
env = dict(os.environ, SWIFTSHEETS_INTEROP_DIR=str(workdir))
failures = []


def expect(cond, what):
    if not cond:
        failures.append(what)


def swift_test(name):
    r = subprocess.run(["swift", "test", "--filter", f"OpenpyxlInteropTests/{name}"], cwd=PKG, env=env, capture_output=True, text=True)
    ok = r.returncode == 0 and "✘" not in r.stdout
    if not ok:
        failures.append(f"swift test {name}:\n{r.stdout[-3000:]}\n{r.stderr[-2000:]}")
    return ok


# ---- 1. SwiftSheets → openpyxl -------------------------------------------------------------------
if swift_test("writesVerificationWorkbook"):
    wb = openpyxl.load_workbook(workdir / "swiftsheets.xlsx")
    ws = wb["Plan"]
    expect(wb.sheetnames == ["Plan", "Hidden"] and wb["Hidden"].sheet_state == "hidden", "sheet list / hidden state")
    expect(wb.properties.creator == "interop" and wb.properties.title == "Interop", "core properties")
    expect(wb.defined_names["PlanRange"].attr_text == "Plan!$A$1:$H$6", "defined name")
    expect(ws["A1"].value == "Title" and ws["A1"].font.b and ws["A1"].font.sz == 14 and ws["A1"].font.color.rgb == "FF112233", "A1 font")
    expect(ws["B1"].value == 42 and isinstance(ws["B1"].value, int) and ws["C1"].value == 3.5 and ws["D1"].value is True, "B1:D1 types")
    expect(ws["E1"].value == dt.datetime(2026, 9, 1) and ws["E1"].number_format == "yyyy/m/d", "E1 date")
    expect(ws["F1"].value == dt.datetime(2026, 9, 1, 13, 30), "F1 datetime")
    expect(ws["G1"].value == "=B1*2", "G1 formula")
    expect(ws["H1"].value == 0.25 and ws["H1"].number_format == "0%", "H1 percent")
    expect(ws["A2"].value == "  padded  ", "A2 whitespace preserved")
    expect(ws["B2"].value == "multi\nline" and ws["B2"].alignment.wrap_text and ws["B2"].alignment.horizontal == "center" and ws["B2"].alignment.vertical == "top", "B2 alignment")
    expect(ws["C2"].fill.patternType == "solid" and ws["C2"].fill.fgColor.rgb == "FFBFD7F5", "C2 fill")
    expect(ws["D2"].border.left.style == "thin" and ws["D2"].border.left.color.rgb == "FF888888" and ws["D2"].border.right.style == "medium", "D2 border")
    expect(ws["E2"].value == '<A&B> "q" 日本語', "E2 escaping")
    expect(ws["F2"].value == "設計 レビュー", "F2 rich text (plain view)")
    expect(ws["G2"].value == dt.time(9, 30) and ws["H2"].value == "#N/A", "G2 time / H2 error")
    expect(str(ws.merged_cells) == "A3:C3" and ws["A3"].value == "merged", "merge")
    expect(ws.freeze_panes == "B2", "freeze panes")
    expect(ws.column_dimensions["A"].width == 20 and ws.column_dimensions["C"].hidden, "column dimensions")
    expect(ws.row_dimensions[2].height == 30 and ws.row_dimensions[4].hidden and ws.row_dimensions[4].outline_level == 1, "row dimensions")
    expect(ws["A6"].hyperlink.target == "https://example.com/", "hyperlink")
    expect(ws.sheet_properties.outlinePr.summaryBelow is False and ws.auto_filter.ref == "A1:H1", "sheet properties / auto filter")
    expect(ws.print_title_rows == "$1:$1" and ws.print_area == "'Plan'!$A$1:$H$6", "print titles / area")
    cols = ws.auto_filter.filterColumn
    expect(len(cols) == 2 and cols[0].colId == 0 and list(cols[0].filters.filter) == ["Title"] and cols[0].filters.blank, "filter values")
    expect(cols[1].colId == 1 and cols[1].customFilters.customFilter[0].operator == "greaterThan"
           and cols[1].customFilters.customFilter[0].val == "10", "custom filter")
    expect(ws.auto_filter.sortState is not None and ws.auto_filter.sortState.sortCondition[0].descending
           and ws.auto_filter.sortState.sortCondition[0].ref in ("B1", "B1:B1"), "sort state")   # a one-cell range writes as "B1"
    dvs = {str(d.sqref): d for d in ws.data_validations.dataValidation}
    lst, whole = dvs.get("C4:C6"), dvs.get("D4:D6")
    expect(lst is not None and lst.type == "list" and lst.formula1 == '"Todo,Doing,Done"'
           and lst.allowBlank and not lst.showErrorMessage and not lst.showDropDown, "list validation (suggests, does not reject)")
    expect(whole is not None and whole.type == "whole" and whole.operator == "between" and whole.formula1 == "0"
           and whole.formula2 == "100" and whole.showErrorMessage and whole.errorStyle == "stop"
           and whole.errorTitle == "範囲外", "whole-number validation")
    expect(ws.oddHeader.left.text == "四半期報告" and ws.oddHeader.center.text == "&P" and ws.oddFooter.right.text == "&F", "header / footer")
    expect([b.id for b in ws.row_breaks.brk] == [4] and [b.id for b in ws.col_breaks.brk] == [2], "page breaks")
    expect(str(ws["B7"].value) == "=SUM(B1:B2)" or getattr(ws["B7"].value, "text", None) == "=SUM(B1:B2)", "array formula value")
    expect(getattr(ws["B7"].value, "ref", None) == "B7:B8", "array formula range")
    expect(ws["A7"].comment is not None and ws["A7"].comment.text == "確認してください\n2 行目" and ws["A7"].comment.author == "南部", "cell note")
    expect([s.name for s in wb._named_styles] == ["Normal", "Accent X"], "named styles declared")
    expect(ws["B6"].style == "Accent X" and ws["B6"].value == 1000, "named style applied to B6")
    accent = wb._named_styles["Accent X"]
    expect(accent.font.b and accent.font.sz == 12 and accent.fill.fgColor.rgb == "FFFFF2CC" and accent.number_format == "#,##0", "named style formatting")

# ---- 1b. SwiftSheets → openpyxl: the features added after the first pass ---------------------------
if swift_test("writesFeatureWorkbook"):
    wb = openpyxl.load_workbook(workdir / "features.xlsx")
    ws = wb["Data"]
    rules = {str(rng): rs for rng, rs in ws.conditional_formatting._cf_rules.items()}
    by_range = {k: v for k, v in rules.items()}
    expect(sum(len(v) for v in by_range.values()) == 6, f"six conditional-formatting rules, got {by_range}")
    cell_is = [r for v in by_range.values() for r in v if r.type == "cellIs"]
    expect(len(cell_is) == 1 and cell_is[0].operator == "greaterThan" and cell_is[0].formula == ["3"], "cellIs rule")
    expect(cell_is[0].dxfId is not None and wb._differential_styles[cell_is[0].dxfId].font.b
           and wb._differential_styles[cell_is[0].dxfId].fill.bgColor.rgb == "FFFFC7CE"
           or wb._differential_styles[cell_is[0].dxfId].fill.fgColor.rgb == "FFFFC7CE", "differential style")
    kinds = sorted({r.type for v in by_range.values() for r in v})
    expect(kinds == ["cellIs", "colorScale", "dataBar", "expression", "iconSet", "top10"], f"rule kinds {kinds}")
    scale = [r for v in by_range.values() for r in v if r.type == "colorScale"][0]
    expect(len(scale.colorScale.cfvo) == 3 and len(scale.colorScale.color) == 3, "colour scale")
    bar = [r for v in by_range.values() for r in v if r.type == "dataBar"][0]
    expect(bar.dataBar.minLength == 10 and bar.dataBar.maxLength == 90, "data bar")
    icons = [r for v in by_range.values() for r in v if r.type == "iconSet"][0]
    expect(icons.iconSet.iconSet == "3Arrows" and len(icons.iconSet.cfvo) == 3, "icon set")
    priorities = sorted(r.priority for v in by_range.values() for r in v)
    expect(priorities == [1, 2, 3, 4, 5, 6], f"priorities renumbered 1..n, got {priorities}")

    expect(list(ws.tables) == ["Sales"], f"named table, got {list(ws.tables)}")
    tbl = ws.tables["Sales"]
    expect(tbl.ref == "A1:C4" and [c.name for c in tbl.tableColumns] == ["Item", "Qty", "Price"], "table columns")
    expect(tbl.tableStyleInfo.name == "TableStyleMedium9" and tbl.tableStyleInfo.showRowStripes, "table style")

    grad = ws["E1"].fill      # a StyleProxy, so check what it says rather than what class it is
    expect(grad.tagname == "gradientFill" and grad.type == "linear" and grad.degree == 90
           and [s.position for s in grad.stop] == [0.0, 1.0]
           and [s.color.rgb for s in grad.stop] == ["FFFFFFFF", "FFBFD7F5"], f"gradient fill, got {grad}")

    top10 = wb["Filtered"].auto_filter.filterColumn[0].top10
    expect(top10 is not None and top10.val == 3 and top10.top is False and top10.percent, f"top-10 filter, got {top10}")

    props = {p.name: p.value for p in wb.custom_doc_props}
    expect(props.get("管理番号") == "A-1234" and props.get("改訂") == 7 and props.get("社外秘") is True,
           f"custom document properties, got {props}")

    pivots = wb["Pivot"]._pivots
    expect(len(pivots) == 1, f"one pivot table, got {len(pivots)}")
    if pivots:
        pt = pivots[0]
        expect(pt.name == "集計" and pt.location.ref == "A3:C5", f"pivot name / location, got {pt.name} {pt.location.ref}")
        expect([f.axis for f in pt.pivotFields] == ["axisRow", None, None], f"pivot field axes, got {[f.axis for f in pt.pivotFields]}")
        expect([f.x for f in (pt.rowFields or [])] == [0], "row field")
        expect([f.x for f in (pt.colFields or [])] == [-2], "the values field is placed across the top")
        expect([(d.fld, d.subtotal) for d in pt.dataFields] == [(1, "sum"), (2, "average")],
               f"data fields, got {[(d.fld, d.subtotal) for d in pt.dataFields]}")
        cache = pt.cache
        expect(cache.cacheSource.worksheetSource.ref == "A1:C4" and cache.cacheSource.worksheetSource.sheet == "Data",
               "cache source")
        expect([f.name for f in cache.cacheFields] == ["Item", "Qty", "Price"], "cache fields")
        expect(cache.refreshOnLoad is True, "cache asks for a refresh on load")
    names = set(openpyxl.load_workbook(workdir / "features.xlsx").sheetnames)
    expect(names == {"Data", "Pivot", "Filtered"}, f"sheets, got {names}")

# ---- 1c. the streaming writer's output, read as an ordinary workbook -------------------------------
if swift_test("writesStreamedWorkbook"):
    wb = openpyxl.load_workbook(workdir / "streamed.xlsx")
    expect(wb.sheetnames == ["Big", "Second"], f"streamed sheets, got {wb.sheetnames}")
    ws = wb["Big"]
    expect(ws.max_row == 2002 and ws["A1"].value == "n" and ws["C2"].value == "行 1", "streamed values")
    expect(ws["A2001"].value == 2000 and ws["B2001"].value == 2000 * 2000, "streamed last data row")
    head = ws["A2002"]
    expect(head.value == "見出し" and head.font.b and head.fill.fgColor.rgb == "FFBFD7F5"
           and head.number_format == "0.00", "streamed styling")
    other = wb["Second"]
    expect(other["A1"].value == "  余白  " and other["B1"].value is True
           and other["C1"].value == "#N/A" and other["D1"].value == 1.5, "streamed types")
    # and openpyxl's own read-only reader walks it too
    ro = openpyxl.load_workbook(workdir / "streamed.xlsx", read_only=True)
    expect(sum(1 for _ in ro["Big"].iter_rows(values_only=True)) == 2002, "openpyxl read_only row count")
    ro.close()

# ---- 2. openpyxl → SwiftSheets -------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Plan"
ws["A1"] = "Title"; ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FF112233")
ws["B1"] = 42; ws["C1"] = 3.5; ws["D1"] = True
ws["E1"] = dt.date(2026, 9, 1); ws["E1"].number_format = "yyyy/m/d"
ws["F1"] = dt.datetime(2026, 9, 1, 13, 30)
ws["G1"] = "=B1*2"; ws["H1"] = 0.25; ws["H1"].number_format = "0%"
ws["A2"] = "  padded  "; ws["B2"] = "multi\nline"; ws["B2"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
ws["C2"].fill = PatternFill("solid", start_color="FFBFD7F5", end_color="FFBFD7F5")
ws["D2"].border = Border(left=Side(style="thin", color="FF888888"), right=Side(style="medium"))
ws["E2"] = '<A&B> "q" 日本語'
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
ws["F2"] = CellRichText(["設計 ", TextBlock(InlineFont(b=True), "レビュー")])
ws["G2"] = dt.time(9, 30); ws["G2"].number_format = "h:mm"; ws["H2"] = "#N/A"
ws["A3"] = "merged"; ws.merge_cells("A3:C3")
ws.freeze_panes = "B2"
ws.column_dimensions["A"].width = 20; ws.column_dimensions["C"].hidden = True
ws.row_dimensions[2].height = 30; ws.row_dimensions[4].hidden = True; ws.row_dimensions[4].outline_level = 1
ws["A4"] = "hidden row"; ws["A5"] = "level 1"
ws["A6"].hyperlink = "https://example.com/"; ws["A6"] = "link"
ws.sheet_properties.outlinePr.summaryBelow = False
ws.auto_filter.ref = "A1:H1"
ws.auto_filter.add_filter_column(0, ["Title"], blank=True)
from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn as PyFilterColumn, SortState as PySortState, SortCondition as PySortCondition
ws.auto_filter.filterColumn.append(PyFilterColumn(colId=1, customFilters=CustomFilters(customFilter=[CustomFilter(operator="greaterThan", val="10")])))
ws.auto_filter.sortState = PySortState(ref="A1:H1", sortCondition=[PySortCondition(ref="B1:B1", descending=True)])
ws.print_title_rows = "1:1"; ws.print_area = "A1:H6"
wb.create_sheet("Hidden").sheet_state = "hidden"; wb["Hidden"]["A1"] = "secret"
wb.properties.creator = "interop"; wb.properties.title = "Interop"
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names["PlanRange"] = DefinedName("PlanRange", attr_text="Plan!$A$1:$H$6")
accent = NamedStyle(name="Accent X")
accent.font = Font(name="Arial", size=12, bold=True, color="FF7F6000")
accent.fill = PatternFill("solid", fgColor="FFFFF2CC")
accent.number_format = "#,##0"
wb.add_named_style(accent)
ws["B6"] = 1000; ws["B6"].style = "Accent X"
from openpyxl.comments import Comment
ws["A7"].comment = Comment("確認してください\n2 行目", "南部")
ws.oddHeader.left.text = "四半期報告"; ws.oddHeader.center.text = "&P"; ws.oddFooter.right.text = "&F"
from openpyxl.worksheet.pagebreak import Break, ColBreak
rb = Break(); rb.id = 4; ws.row_breaks.append(rb)
cb = ColBreak(); ws.col_breaks.append(cb); ws.col_breaks.brk[0].id = 2   # append renumbers, so set it after
from openpyxl.worksheet.formula import ArrayFormula
ws["B7"] = ArrayFormula("B7:B8", "=SUM(B1:B2)")
from openpyxl.worksheet.datavalidation import DataValidation as PyDataValidation
dv = PyDataValidation(type="list", formula1='"Todo,Doing,Done"', allow_blank=True, showErrorMessage=False)
ws.add_data_validation(dv); dv.add("C4:C6")
wb.save(workdir / "openpyxl.xlsx")
swift_test("readsVerificationWorkbook")

result = {"openpyxl": openpyxl.__version__, "swiftsheets_to_openpyxl": all("swift test writesVerificationWorkbook" not in f for f in failures) and not [f for f in failures if not f.startswith("swift test")],
          "openpyxl_to_swiftsheets": not [f for f in failures if f.startswith("swift test readsVerificationWorkbook")], "failures": failures, "workdir": str(workdir)}
print(json.dumps(result, ensure_ascii=False, indent=1))
sys.exit(1 if failures else 0)
