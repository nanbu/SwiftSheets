"""Generates Tests/SwiftSheetsTests/Fixtures/*.xlsx with openpyxl (the behavioural reference) — run with `uv run` or any
Python that has openpyxl. Fixtures are committed; regenerate only when the reference behaviour matters."""
import datetime as dt
import io
import pathlib
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils.datetime import CALENDAR_MAC_1904

OUT = pathlib.Path(__file__).resolve().parents[1] / "SwiftSheetsTests" / "Fixtures"
OUT.mkdir(parents=True, exist_ok=True)


def styled():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Title"; ws["A1"].font = Font(name="Arial", size=14, bold=True, italic=True, color="FF112233")
    ws["B1"] = 42; ws["C1"] = 3.5; ws["D1"] = True; ws["E1"] = dt.date(2026, 9, 1); ws["E1"].number_format = "yyyy/m/d"
    ws["F1"] = dt.datetime(2026, 9, 1, 13, 30); ws["F1"].number_format = "yyyy-mm-dd hh:mm"
    ws["G1"] = "=B1*2"; ws["H1"] = 0.25; ws["H1"].number_format = "0%"
    ws["A2"] = "  padded  "; ws["B2"] = "multi\nline"; ws["B2"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
    ws["C2"].fill = PatternFill("solid", start_color="FFBFD7F5", end_color="FFBFD7F5")
    thin = Side(style="thin", color="FF888888")
    ws["D2"].border = Border(left=thin, right=Side(style="medium"), top=thin, bottom=thin)
    ws["E2"] = "<A&B> \"q\""
    ws.merge_cells("A3:C3"); ws["A3"] = "merged"
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["C"].hidden = True
    ws.row_dimensions[2].height = 30; ws.row_dimensions[4].hidden = True; ws.row_dimensions[4].outline_level = 1
    ws.row_dimensions[5].outline_level = 1; ws.row_dimensions[5].collapsed = True
    ws["A4"] = "hidden row"; ws["A5"] = "level 1"
    ws["A6"].hyperlink = "https://example.com/"; ws["A6"] = "link"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.auto_filter.ref = "A1:H1"
    wb.create_sheet("Hidden").sheet_state = "hidden"
    wb["Hidden"]["A1"] = "secret"
    wb.properties.creator = "fixture"
    wb.properties.title = "Styled"
    wb.save(OUT / "styled.xlsx")


def date1904():
    wb = Workbook(); wb.epoch = CALENDAR_MAC_1904
    ws = wb.active; ws["A1"] = dt.date(2026, 9, 1); ws["A1"].number_format = "yyyy/m/d"; ws["B1"] = 7
    wb.save(OUT / "date1904.xlsx")


def rph():
    """Hand-built OOXML with furigana runs and rich text in sharedStrings (Japanese Excel), stored entries."""
    ct = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/></Types>'
    rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'
    wb = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="工程表" sheetId="1" r:id="rId1"/></sheets></workbook>'
    wbrels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>'
    sst = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="3" uniqueCount="3"><si><t>要件定義</t><rPh sb="0" eb="4"><t>ヨウケンテイギ</t></rPh><phoneticPr fontId="1"/></si><si><r><t xml:space="preserve">設計 </t></r><r><rPr><b/><sz val="12"/></rPr><t>レビュー</t></r><rPh sb="0" eb="2"><t>セッケイ</t></rPh></si><si><t xml:space="preserve">　字下げ</t></si></sst>'
    styles = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="1"><font><sz val="11"/><name val="游ゴシック"/></font></fonts><fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="14" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/></cellXfs></styleSheet>'
    sheet = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData><row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" s="1"><v>46266</v></c><c t="inlineStr"><is><t>inline</t></is></c><c r="D1" t="str"><f>A1</f><v>要件定義</v></c></row><row r="2" outlineLevel="1" hidden="1"><c r="A2" t="s"><v>1</v></c><c r="B2"><v>2.0</v></c><c r="C2" t="b"><v>1</v></c><c r="D2" t="e"><v>#DIV/0!</v></c></row><row r="3"><c r="A3" t="s"><v>2</v></c><c r="B3" s="1"><v>0.5</v></c></row></sheetData></worksheet>'
    with zipfile.ZipFile(OUT / "rph.xlsx", "w", zipfile.ZIP_STORED) as z:
        z.writestr("[Content_Types].xml", ct); z.writestr("_rels/.rels", rels); z.writestr("xl/workbook.xml", wb)
        z.writestr("xl/_rels/workbook.xml.rels", wbrels); z.writestr("xl/sharedStrings.xml", sst); z.writestr("xl/styles.xml", styles)
        z.writestr("xl/worksheets/sheet1.xml", sheet)


def named_styles():
    """Named cell styles: one of Excel's builtins, one of the file's own, and a cell that keeps the link while
    overriding part of the formatting (openpyxl writes the override into cellXfs and leaves xfId pointing home)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Styles"
    heading = NamedStyle(name="Heading X")
    heading.font = Font(bold=True, size=14, color="FF1F4E79")
    heading.fill = PatternFill("solid", fgColor="FFDDEBF7")
    heading.border = Border(bottom=Side(style="thin"))
    heading.alignment = Alignment(horizontal="center")
    heading.number_format = "0.00"
    wb.add_named_style(heading)
    ws["A1"] = "見出し"; ws["A1"].style = "Heading X"
    ws["A2"] = 1; ws["A2"].style = "Heading X"; ws["A2"].font = Font(italic=True)   # link kept, font overridden
    ws["A3"] = 2                                                                     # Normal
    ws["B1"] = 3; ws["B1"].style = "Title"                                           # a builtin
    wb.save(OUT / "named-styles.xlsx")


styled(); date1904(); rph(); named_styles()
print("fixtures written to", OUT)
