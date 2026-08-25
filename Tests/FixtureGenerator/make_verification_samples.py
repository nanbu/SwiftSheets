# -*- coding: utf-8 -*-
"""Builds the seed workbook for the manual verification samples (MAINTENANCE.md's release checklist): realistic
Japanese business data exercising formatting, colours, borders, filters, formulas, notes, links, hidden sheets and
multi-byte text. LibreOffice then converts it to ODF, and that LibreOffice-written .ods is what SwiftSheets reads.

    python3 Tests/FixtureGenerator/make_verification_samples.py <out.xlsx>     # needs openpyxl
"""
import datetime as dt
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

NAVY = "FF1F3864"
BLUE = "FFD9E2F3"
GREY = "FFF2F2F2"
GREEN = "FFE2EFDA"
RED = "FFFCE4EC"
YELLOW = "FFFFF2CC"

thin = Side(style="thin", color="FF9DA5B4")
medium = Side(style="medium", color="FF1F3864")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

YEN = '"¥"#,##0'
YEN2 = '"¥"#,##0.00'
DATE_JP = 'yyyy"年"m"月"d"日"'

SALES = [
    ("2026-04-03", "株式会社山田製作所", "第一営業部", "佐藤 花子", "精密ベアリング A-100", 120, 3800, 0.284),
    ("2026-04-11", "中村商事株式会社", "第一営業部", "佐藤 花子", "精密ベアリング B-220", 45, 12500, 0.312),
    ("2026-04-18", "有限会社小林工業", "第二営業部", "鈴木 一郎", "駆動ユニット DX-5", 8, 148000, 0.221),
    ("2026-05-07", "株式会社山田製作所", "第一営業部", "田中 美咲", "精密ベアリング A-100", 260, 3650, 0.268),
    ("2026-05-15", "グローバル電機株式会社", "海外営業部", "John Smith", "制御基板 CTL-9", 75, 22400, 0.355),
    ("2026-05-22", "中村商事株式会社", "第一営業部", "佐藤 花子", "駆動ユニット DX-5", 12, 145000, 0.198),
    ("2026-06-02", "株式会社青木エンジニアリング", "第二営業部", "鈴木 一郎", "精密ベアリング B-220", 90, 12100, 0.301),
    ("2026-06-09", "有限会社小林工業", "第二営業部", "高橋 健", "制御基板 CTL-9", 30, 23000, 0.340),
    ("2026-06-19", "グローバル電機株式会社", "海外営業部", "John Smith", "駆動ユニット DX-5", 20, 142000, 0.205),
    ("2026-07-01", "株式会社山田製作所", "第一営業部", "田中 美咲", "制御基板 CTL-9", 55, 22800, 0.348),
    ("2026-07-14", "株式会社青木エンジニアリング", "第二営業部", "高橋 健", "精密ベアリング A-100", 310, 3550, 0.259),
    ("2026-07-28", "中村商事株式会社", "第一営業部", "佐藤 花子", "精密ベアリング B-220", 66, 12300, 0.295),
    ("2026-08-05", "有限会社小林工業", "第二営業部", "鈴木 一郎", "駆動ユニット DX-5", 6, 150000, 0.230),
    ("2026-08-12", "グローバル電機株式会社", "海外営業部", "John Smith", "制御基板 CTL-9", 95, 21900, 0.362),
    ("2026-08-26", "株式会社山田製作所", "第一営業部", "田中 美咲", "精密ベアリング A-100", 180, 3700, 0.275),
    ("2026-09-04", "株式会社青木エンジニアリング", "第二営業部", "高橋 健", "駆動ユニット DX-5", 15, 146500, 0.212),
    ("2026-09-17", "中村商事株式会社", "第一営業部", "佐藤 花子", "制御基板 CTL-9", 40, 22600, 0.351),
    ("2026-09-30", "有限会社小林工業", "第二営業部", "鈴木 一郎", "精密ベアリング B-220", 120, 11900, 0.288),
]

STAFF = [
    ("0012", "佐藤 花子", "サトウ ハナコ", "第一営業部", "2018-04-01", "内線 1204", "hanako.sato@example.co.jp", True,
     "産休から復帰（2026-04）。\n主要顧客: 山田製作所・中村商事"),
    ("0027", "鈴木 一郎", "スズキ イチロウ", "第二営業部", "2015-10-01", "内線 1318", "ichiro.suzuki@example.co.jp", True,
     "半導体装置分野の担当。TOEIC 820"),
    ("0104", "田中 美咲", "タナカ ミサキ", "第一営業部", "2022-04-01", "内線 1210", "misaki.tanaka@example.co.jp", True,
     "新人研修 修了。ﾊﾝｶｸｶﾅのテスト"),
    ("0140", "高橋 健", "タカハシ ケン", "第二営業部", "2024-04-01", "内線 1325", "ken.takahashi@example.co.jp", True, ""),
    ("0009", "John Smith", "ジョン スミス", "海外営業部", "2012-06-15", "内線 1401", "john.smith@example.co.jp", False,
     "2026-03 に休職。復帰未定"),
]


def money(ws, cell, fmt=YEN):
    ws[cell].number_format = fmt


def build(path):
    wb = Workbook()

    # ---------------- Sheet 1: 売上明細 ----------------
    ws = wb.active
    ws.title = "Sales"
    ws.merge_cells("A1:I1")
    ws["A1"] = "2026年度 上期 売上明細"
    ws["A1"].font = Font(name="游ゴシック", size=16, bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A2"] = "作成日"
    ws["B2"] = dt.date(2026, 10, 1)
    ws["B2"].number_format = DATE_JP
    ws["D2"] = "単位: 円（税抜）"
    ws["D2"].font = Font(italic=True, color="FF7F7F7F")
    ws["H2"] = "社外秘"
    ws["H2"].font = Font(bold=True, color="FFC00000")
    ws["H2"].alignment = Alignment(horizontal="right")

    headers = ["日付", "取引先", "部門", "担当者", "商品名", "数量", "単価", "金額", "粗利率"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box
    ws.row_dimensions[3].height = 22

    for n, (date, customer, dept, owner, product, qty, price, margin) in enumerate(SALES):
        r = 4 + n
        ws.cell(row=r, column=1, value=dt.date.fromisoformat(date)).number_format = "yyyy/m/d"
        ws.cell(row=r, column=2, value=customer)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=owner)
        ws.cell(row=r, column=5, value=product)
        ws.cell(row=r, column=6, value=qty).number_format = "#,##0"
        ws.cell(row=r, column=7, value=price).number_format = YEN
        ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = YEN
        ws.cell(row=r, column=9, value=margin).number_format = "0.0%"
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            if n % 2 == 1:
                cell.fill = PatternFill("solid", start_color=GREY, end_color=GREY)
        if margin < 0.22:                                   # thin margin: red text, pink fill
            ws.cell(row=r, column=9).font = Font(bold=True, color="FFC00000")
            ws.cell(row=r, column=9).fill = PatternFill("solid", start_color=RED, end_color=RED)
        elif margin > 0.34:                                 # good margin: green
            ws.cell(row=r, column=9).font = Font(bold=True, color="FF1E7145")
            ws.cell(row=r, column=9).fill = PatternFill("solid", start_color=GREEN, end_color=GREEN)

    last = 3 + len(SALES)
    total = last + 1
    ws.cell(row=total, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total, column=6, value=f"=SUM(F4:F{last})").number_format = "#,##0"
    ws.cell(row=total, column=8, value=f"=SUM(H4:H{last})").number_format = YEN
    ws.cell(row=total, column=9, value=f"=AVERAGE(I4:I{last})").number_format = "0.0%"
    for c in range(1, 10):
        cell = ws.cell(row=total, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color=BLUE, end_color=BLUE)
        cell.border = Border(left=thin, right=thin, top=medium, bottom=Side(style="double", color=NAVY))

    ws.cell(row=total + 2, column=1, value="備考: 金額は数量×単価。粗利率 22% 未満は赤、34% 超は緑。")
    ws.cell(row=total + 2, column=1).font = Font(size=9, color="FF7F7F7F")
    ws["B4"].comment = Comment("与信枠 3,000 万円。支払サイト 45 日。", "経理部")
    ws["E2"] = "社内Wiki"
    ws["E2"].hyperlink = "https://example.co.jp/wiki/sales"
    ws["E2"].font = Font(color="FF0563C1", underline="single")

    widths = {"A": 12, "B": 30, "C": 14, "D": 14, "E": 26, "F": 8, "G": 12, "H": 14, "I": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{last}"

    # ---------------- Sheet 2: 月次サマリ ----------------
    s2 = wb.create_sheet("Monthly")
    s2["A1"] = "月次サマリ（売上明細シートを集計）"
    s2["A1"].font = Font(size=14, bold=True, color=NAVY)
    s2.merge_cells("A1:E1")
    months = ["4月", "5月", "6月", "7月", "8月", "9月"]
    for i, h in enumerate(["月", "売上", "予算", "達成率", "判定"], start=1):
        c = s2.cell(row=3, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color=BLUE, end_color=BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = box
    budgets = [1_500_000, 3_000_000, 3_200_000, 3_000_000, 3_400_000, 3_100_000]
    for i, (m, budget) in enumerate(zip(months, budgets)):
        r = 4 + i
        s2.cell(row=r, column=1, value=m).alignment = Alignment(horizontal="center")
        s2.cell(row=r, column=2, value=f'=SUMPRODUCT((TEXT(Sales!$A$4:$A${last},"m")="{m[:-1]}")*Sales!$H$4:$H${last})').number_format = YEN
        s2.cell(row=r, column=3, value=budget).number_format = YEN
        s2.cell(row=r, column=4, value=f"=B{r}/C{r}").number_format = "0.0%"
        s2.cell(row=r, column=5, value=f'=IF(B{r}>=C{r},"達成","未達")').alignment = Alignment(horizontal="center")
        for c in range(1, 6):
            s2.cell(row=r, column=c).border = box
    s2.cell(row=10, column=1, value="合計").font = Font(bold=True)
    s2.cell(row=10, column=2, value="=SUM(B4:B9)").number_format = YEN
    s2.cell(row=10, column=3, value="=SUM(C4:C9)").number_format = YEN
    s2.cell(row=10, column=4, value="=B10/C10").number_format = "0.0%"
    for c in range(1, 6):
        cell = s2.cell(row=10, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color=YELLOW, end_color=YELLOW)
        cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    for col, w in {"A": 10, "B": 16, "C": 16, "D": 12, "E": 10}.items():
        s2.column_dimensions[col].width = w
    s2.freeze_panes = "A4"

    # ---------------- Sheet 3: 社員名簿 ----------------
    s3 = wb.create_sheet("Staff")
    for i, h in enumerate(["社員番号", "氏名", "フリガナ", "部署", "入社日", "連絡先", "メール", "在籍", "備考"], start=1):
        c = s3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color="FF375623", end_color="FF375623")
        c.alignment = Alignment(horizontal="center")
        c.border = box
    for n, (num, name, kana, dept, joined, ext, mail, active, note) in enumerate(STAFF):
        r = 2 + n
        s3.cell(row=r, column=1, value=num).alignment = Alignment(horizontal="center")   # text: keeps the leading zero
        s3.cell(row=r, column=2, value=name)
        s3.cell(row=r, column=3, value=kana).font = Font(size=9)
        s3.cell(row=r, column=4, value=dept)
        s3.cell(row=r, column=5, value=dt.date.fromisoformat(joined)).number_format = "yyyy/mm/dd"
        s3.cell(row=r, column=6, value=ext)
        cell = s3.cell(row=r, column=7, value=mail)
        cell.hyperlink = "mailto:" + mail
        cell.font = Font(color="FF0563C1", underline="single")
        s3.cell(row=r, column=8, value=active).alignment = Alignment(horizontal="center")
        s3.cell(row=r, column=9, value=note).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 10):
            s3.cell(row=r, column=c).border = box
        if not active:
            for c in range(1, 10):
                s3.cell(row=r, column=c).font = Font(color="FF808080", italic=True)
        s3.row_dimensions[r].height = 30
    for col, w in {"A": 10, "B": 16, "C": 18, "D": 14, "E": 12, "F": 12, "G": 30, "H": 8, "I": 34}.items():
        s3.column_dimensions[col].width = w
    s3.freeze_panes = "B2"

    # ---------------- Sheet 4: 書式見本 ----------------
    s4 = wb.create_sheet("Formats")
    s4["A1"] = "書式・表示形式・罫線の見本"
    s4["A1"].font = Font(size=14, bold=True)
    s4.merge_cells("A1:D1")

    s4["A3"] = "表示形式"
    s4["A3"].font = Font(bold=True)
    formats = [
        ("General", 1234.5678, "General"), ("0", 1234.5678, "0"), ("0.00", 1234.5678, "0.00"),
        ("#,##0", 1234567, "#,##0"), ('"¥"#,##0', 1234567, YEN), ('"¥"#,##0.00', 1234.5, YEN2),
        ("0%", 0.1234, "0%"), ("0.0%", 0.1234, "0.0%"),
        ("yyyy/m/d", dt.date(2026, 9, 1), "yyyy/m/d"), ('yyyy"年"m"月"d"日"', dt.date(2026, 9, 1), DATE_JP),
        ("h:mm", dt.time(9, 30), "h:mm"), ("h:mm:ss", dt.time(13, 5, 30), "h:mm:ss"),
        ("[h]:mm:ss（経過時間）", dt.timedelta(hours=26, minutes=1, seconds=30), "[h]:mm:ss"),
        ("@（文字列）", "0012-345", "@"),
    ]
    for i, (label, value, fmt) in enumerate(formats):
        r = 4 + i
        s4.cell(row=r, column=1, value=label).font = Font(size=10)
        c = s4.cell(row=r, column=2, value=value)
        c.number_format = fmt
        s4.cell(row=r, column=3, value=fmt).font = Font(name="Menlo", size=9, color="FF7F7F7F")

    s4["E3"] = "罫線"
    s4["E3"].font = Font(bold=True)
    for i, style in enumerate(["thin", "medium", "thick", "double", "dashed", "dotted", "hair"]):
        r = 4 + i
        s4.cell(row=r, column=5, value=style).font = Font(size=10)
        side = Side(style=style, color="FF1F3864")
        s4.cell(row=r, column=6, value="サンプル").border = Border(left=side, right=side, top=side, bottom=side)

    s4["E12"] = "塗りつぶし"
    s4["E12"].font = Font(bold=True)
    for i, (label, colour) in enumerate([("薄い青", BLUE), ("薄い緑", GREEN), ("薄い赤", RED), ("薄い黄", YELLOW), ("灰", GREY), ("濃紺（白文字）", NAVY)]):
        r = 13 + i
        s4.cell(row=r, column=5, value=label).font = Font(size=10)
        c = s4.cell(row=r, column=6, value="■■■")
        c.fill = PatternFill("solid", start_color=colour, end_color=colour)
        if colour == NAVY:
            c.font = Font(color="FFFFFFFF", bold=True)

    s4["A20"] = "配置とフォント"
    s4["A20"].font = Font(bold=True)
    samples = [
        ("左寄せ", Alignment(horizontal="left"), Font()),
        ("中央", Alignment(horizontal="center"), Font()),
        ("右寄せ", Alignment(horizontal="right"), Font()),
        ("上詰め", Alignment(vertical="top"), Font()),
        ("下詰め", Alignment(vertical="bottom"), Font()),
        ("折り返して全体を表示する長い日本語のテキストです。", Alignment(wrap_text=True, vertical="top"), Font(size=10)),
        ("太字", Alignment(), Font(bold=True)),
        ("斜体", Alignment(), Font(italic=True)),
        ("下線", Alignment(), Font(underline="single")),
        ("取り消し線", Alignment(), Font(strikethrough=True)),
        ("16pt 赤", Alignment(), Font(size=16, color="FFC00000")),
        ("游ゴシック", Alignment(), Font(name="游ゴシック", size=12)),
        ("Menlo 等幅", Alignment(), Font(name="Menlo", size=11)),
    ]
    for i, (label, align, font) in enumerate(samples):
        r = 21 + i
        s4.cell(row=r, column=1, value=label).font = Font(size=10)
        c = s4.cell(row=r, column=2, value=label)
        c.alignment = align
        c.font = font
        if "折り返し" in label:
            s4.row_dimensions[r].height = 46

    s4["E20"] = "文字と記号"
    s4["E20"].font = Font(bold=True)
    for i, text in enumerate(["漢字・ひらがな・カタカナ", "ﾊﾝｶｸｶﾅ ABC 123", "①②③ ㈱ ℡ ㎡ №", "𠮷野家（サロゲートペア）", "🎌 🗾 ✅（絵文字）",
                              "改行を\n含む文字列", "  前後の空白  ", "引用符 \" と & < >"]):
        r = 21 + i
        s4.cell(row=r, column=5, value=text).alignment = Alignment(wrap_text="改行" in text, vertical="top")
    for col, w in {"A": 26, "B": 24, "C": 18, "E": 24, "F": 16}.items():
        s4.column_dimensions[col].width = w

    s4["A36"] = "結合セル（A36:C37）"
    s4.merge_cells("A36:C37")
    s4["A36"].alignment = Alignment(horizontal="center", vertical="center")
    s4["A36"].fill = PatternFill("solid", start_color=YELLOW, end_color=YELLOW)
    s4["A36"].border = box

    # ---------------- Sheet 5: 非表示（hidden） ----------------
    s5 = wb.create_sheet("Extra")
    s5["A1"] = "このシートは非表示です（Excel: 右クリック→再表示／Numbers: 非表示シートは表示されます）"
    s5["A2"] = "レート"
    s5["B2"] = 152.35
    s5["B2"].number_format = "0.00"
    s5.column_dimensions["A"].width = 60
    s5.sheet_state = "hidden"

    wb.defined_names["SalesRange"] = DefinedName("SalesRange", attr_text=f"Sales!$A$3:$I${last}")
    wb.properties.creator = "SwiftSheets 検証サンプル"
    wb.properties.title = "SwiftSheets conversion verification sample"
    wb.active = 0
    wb.save(path)
    print(f"wrote {path}: sheets={wb.sheetnames}")


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "verification-seed.xlsx")
