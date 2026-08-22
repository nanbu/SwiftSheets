"""Fixtures for the "we do not open this, and we say why" tests (spec §1.3 / §14.11): a password-protected OOXML
package, a password-protected ODF package, and a legacy .xls. None of them is a SwiftSheets output — the point is
that real encryptors produce them. Regenerate with:

    python3 Tests/FixtureGenerator/make_encrypted_fixtures.py     # needs openpyxl + msoffcrypto-tool; LibreOffice for the last two

The password on all of them is "swiftsheets"; nothing inside matters, because nothing inside is ever read.
"""
import pathlib
import shutil
import subprocess
import sys
import tempfile

from openpyxl import Workbook

OUT = pathlib.Path(__file__).resolve().parents[1] / "SwiftSheetsTests" / "Fixtures" / "encrypted"
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
PASSWORD = "swiftsheets"


def plain_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Secret"
    ws["A1"] = "こんにちは"
    ws["A2"] = 42
    wb.save(path)


def encrypt_odf(source, destination, password):
    """Re-package an ODF document with every entry encrypted, as ODF 1.3 §4.3 prescribes: deflate the entry, then
    AES-256-CBC it under a key derived with PBKDF2-HMAC-SHA1 from SHA-256 of the password, and record the algorithm,
    salt, IV and a SHA-256 checksum of the first 1024 compressed bytes in META-INF/manifest.xml. `mimetype` stays in
    the clear (that is what keeps the file recognisable as a spreadsheet) and so does the manifest itself.
    """
    import base64
    import hashlib
    import os
    import re
    import zipfile
    import zlib

    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes

    ITERATIONS = 100_000
    start_key = hashlib.sha256(password.encode("utf-8")).digest()
    entries, encryption = {}, {}
    with zipfile.ZipFile(source) as zin:
        names = zin.namelist()
        for name in names:
            data = zin.read(name)
            if name == "mimetype" or name.startswith("META-INF/") or name.endswith("/"):
                entries[name] = (data, zipfile.ZIP_STORED)
                continue
            compressed = zlib.compressobj(9, zlib.DEFLATED, -zlib.MAX_WBITS)
            body = compressed.compress(data) + compressed.flush()
            checksum = base64.b64encode(hashlib.sha256(body[:1024]).digest()).decode()
            salt, iv = os.urandom(16), os.urandom(16)
            key = PBKDF2HMAC(algorithm=hashes.SHA1(), length=32, salt=salt, iterations=ITERATIONS).derive(start_key)
            pad = 16 - len(body) % 16
            encryptor = Cipher(algorithms.AES(key), modes.CBC(iv)).encryptor()
            entries[name] = (encryptor.update(body + bytes([pad]) * pad) + encryptor.finalize(), zipfile.ZIP_STORED)
            encryption[name] = (len(data), checksum, base64.b64encode(iv).decode(), base64.b64encode(salt).decode())

    manifest = zipfile.ZipFile(source).read("META-INF/manifest.xml").decode("utf-8")

    def annotate(match):
        path = match.group("path")
        if path not in encryption:
            return match.group(0)
        size, checksum, iv, salt = encryption[path]
        entry = match.group(0)[:-2] + f' manifest:size="{size}">'
        return entry + (
            f'<manifest:encryption-data manifest:checksum-type='
            f'"urn:oasis:names:tc:opendocument:xmlns:manifest:1.0#sha256-1k" manifest:checksum="{checksum}">'
            f'<manifest:algorithm manifest:algorithm-name="http://www.w3.org/2001/04/xmlenc#aes256-cbc"'
            f' manifest:initialisation-vector="{iv}"/>'
            f'<manifest:key-derivation manifest:key-derivation-name="PBKDF2" manifest:key-size="32"'
            f' manifest:iteration-count="{ITERATIONS}" manifest:salt="{salt}"/>'
            f'<manifest:start-key-generation manifest:start-key-generation-name='
            f'"http://www.w3.org/2000/09/xmldsig#sha256" manifest:key-size="32"/>'
            f'</manifest:encryption-data></manifest:file-entry>'
        )

    manifest = re.sub(r'<manifest:file-entry[^>]*manifest:full-path="(?P<path>[^"]+)"[^>]*/>', annotate, manifest)
    entries["META-INF/manifest.xml"] = (manifest.encode("utf-8"), zipfile.ZIP_STORED)

    with zipfile.ZipFile(destination, "w") as zout:
        # mimetype first and stored, as ODF requires of every package, encrypted or not
        zout.writestr("mimetype", entries.pop("mimetype")[0], zipfile.ZIP_STORED)
        for name, (data, method) in entries.items():
            zout.writestr(name, data, method)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    tmp = pathlib.Path(tempfile.mkdtemp())
    plain = tmp / "plain.xlsx"
    plain_workbook(plain)

    # 1. ECMA-376 agile encryption, written by msoffcrypto-tool: an OLE compound file whose directory names an
    #    EncryptedPackage stream.
    from msoffcrypto.format.ooxml import OOXMLFile

    with open(plain, "rb") as f, open(OUT / "agile.xlsx", "wb") as out:
        OOXMLFile(f).encrypt(PASSWORD, out)

    if not pathlib.Path(SOFFICE).exists():
        print("LibreOffice not found — skipping the .ods and .xls fixtures", file=sys.stderr)
        return

    # 2. A password-protected ODF package. LibreOffice's --convert-to ignores a Password filter option (it only
    #    honours one when *reading*), so the package is encrypted here, to ODF 1.3 §4.3 — and LibreOffice is the
    #    judge that it worked: it must refuse to convert the result without the password.
    subprocess.run([SOFFICE, "--headless", "--convert-to", "ods", "--outdir", str(tmp), str(plain)],
                   check=True, capture_output=True)
    encrypt_odf(tmp / "plain.ods", OUT / "protected.ods", PASSWORD)

    # 3. A legacy .xls (BIFF8) — also an OLE compound file, but with no EncryptedPackage stream.
    subprocess.run([SOFFICE, "--headless", "--convert-to", "xls", "--outdir", str(tmp), str(plain)],
                   check=True, capture_output=True)
    shutil.copy(tmp / "plain.xls", OUT / "legacy.xls")

    for p in sorted(OUT.iterdir()):
        print(p.name, p.stat().st_size, "bytes")


if __name__ == "__main__":
    main()
