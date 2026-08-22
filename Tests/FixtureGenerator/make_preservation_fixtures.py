"""Fixtures for the round-trip preservation tests (spec chapter 6 / 12): workbooks carrying parts SwiftSheets does
not interpret — a chart, a comment, a table, conditional formatting, data validation, a defined name, a second sheet
with a drawing — plus a hand-made .xlsm with a (dummy) VBA project. Regenerate with:

    uv run --project ../../../Stream/web python Tests/FixtureGenerator/make_preservation_fixtures.py
"""
import pathlib
import shutil
import zipfile

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = pathlib.Path(__file__).resolve().parents[1] / "SwiftSheetsTests" / "Fixtures" / "preservation"
OUT.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.append(["Item", "Qty", "Price"])
for row in [("apple", 3, 1.5), ("pear", 5, 2.25), ("plum", 2, 4.0)]:
    ws.append(row)
ws["E1"] = "=SUM(B2:B4)"
ws["A1"].font = Font(bold=True)
ws["A1"].comment = Comment("first column", "tester")
chart = BarChart()
chart.title = "Quantities"
chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
ws.add_chart(chart, "G2")
tab = Table(displayName="Items", ref="A1:C4")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws.add_table(tab)
ws.conditional_formatting.add("B2:B4", CellIsRule(operator="greaterThan", formula=["3"], fill=PatternFill("solid", fgColor="FFC7CE")))
dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("D2:D4")
wb.defined_names["Prices"] = DefinedName("Prices", attr_text="Data!$C$2:$C$4")
ws2 = wb.create_sheet("Notes")
ws2["A1"] = "second sheet"
ws2["A2"].comment = Comment("note on sheet two", "tester")
wb.save(OUT / "charts-and-friends.xlsx")

# a macro-enabled copy: same package, macro content type, dummy vbaProject.bin wired in through the workbook rels
src = zipfile.ZipFile(OUT / "charts-and-friends.xlsx")
with zipfile.ZipFile(OUT / "with-vba.xlsm", "w", zipfile.ZIP_DEFLATED) as dst:
    for item in src.infolist():
        data = src.read(item.filename)
        if item.filename == "[Content_Types].xml":
            data = data.replace(b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                                b"application/vnd.ms-excel.sheet.macroEnabled.main+xml")
            data = data.replace(b"</Types>", b'<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>')
        if item.filename == "xl/_rels/workbook.xml.rels":
            data = data.replace(b"</Relationships>", b'<Relationship Id="rId99" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/></Relationships>')
        dst.writestr(item, data)
    dst.writestr("xl/vbaProject.bin", bytes(range(256)) * 4)   # not a real project; the bytes only have to survive untouched
print("wrote", sorted(p.name for p in OUT.iterdir()))
